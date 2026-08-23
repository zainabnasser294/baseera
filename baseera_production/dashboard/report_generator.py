import io
import re
import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime


def sanitize_sheet_title(title: str) -> str:
    """إزالة أي رموز ممنوعة في أسماء ورقات العمل في إكسل وحصر الطول إلى 30 حرفاً."""
    cleaned = re.sub(r'[:\\/\?\*\[\]]', ' ', str(title))
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned[:30] if cleaned else "تقرير بصيرة"


def find_num_val(rdata: dict, keywords: list, default_val=0.0) -> float:
    """استخراج القيمة الرقمية ذكياً من أي قاموس بياني بناءً على الكلمات المفتاحية."""
    if not isinstance(rdata, dict):
        return float(default_val)
    for key, val in rdata.items():
        key_str = str(key).lower().strip()
        for kw in keywords:
            if kw.lower() in key_str:
                try:
                    num_str = str(val).replace(',', '').replace('ر.ع.', '').replace('OMR', '').strip()
                    num = float(num_str)
                    if not pd.isna(num):
                        return num
                except Exception:
                    pass
    return float(default_val)


def find_str_val(rdata: dict, keywords: list, default_val="") -> str:
    """استخراج القيمة النصية ذكياً من أي قاموس بياني بناءً على الكلمات المفتاحية."""
    if not isinstance(rdata, dict):
        return default_val
    for key, val in rdata.items():
        key_str = str(key).lower().strip()
        for kw in keywords:
            if kw.lower() in key_str:
                val_str = str(val).strip()
                if val_str and val_str.lower() != 'nan':
                    return val_str
    if not default_val:
        for key, val in rdata.items():
            if isinstance(val, str) and len(val.strip()) > 1 and val.strip().lower() != 'nan':
                return val.strip()
    return default_val


def generate_baseera_excel(
    user_language="AR", client_data=None, output_target="تقرير_بصيرة.xlsx", output_filename=None
):
    """
    توليد تقرير إكسل "تقرير بصيرة" أنيق وملموم ومصمم باحترافية عالية.
    يأتي التقرير بتنسيق فاخر ومدمج في مكان واحد بدون توسع أفقي زائد.
    """
    if output_filename is not None:
        output_target = output_filename
    lang = user_language.upper() if str(user_language).upper() in ["AR", "EN"] else "AR"

    # استخراج البيانات العامة من الحمولة Payload
    payload = client_data if isinstance(client_data, dict) else {}
    company_name = payload.get("company_name", payload.get("client_name", "المؤسسة / الشركة"))
    currency = payload.get("currency", "ر.ع." if lang == "AR" else "OMR")

    # استخراج الأصناف/الحركات
    raw_items = payload.get("items", payload.get("rows", []))
    if not raw_items and isinstance(client_data, list):
        raw_items = client_data

    items = []
    kw_code = ["كود", "رمز", "معرف", "code", "sku", "id", "ticket", "inv"]
    kw_name = ["اسم", "منتج", "سلعة", "صنف", "بند", "name", "title", "product", "item", "description"]
    kw_cat = ["فئة", "تصنيف", "قسم", "نوع", "category", "type", "dept", "group"]
    kw_qty = ["كمية مباعة", "كمية", "عدد", "qty_sold", "qty", "sold", "quantity", "orders"]
    kw_price = ["سعر البيع", "سعر الوحدة", "سعر", "مبلغ", "إيراد", "unit_price", "price", "rate", "revenue"]
    kw_cost = ["تكلفة الوحدة", "تكلفة", "مصروف", "unit_cost", "cost", "cogs"]
    kw_wasted = ["كمية مهدرة", "هدر", "خسارة", "تالف", "qty_wasted", "waste", "loss"]

    if raw_items:
        for idx, item_dict in enumerate(raw_items):
            code = find_str_val(item_dict, kw_code, f"SKU-{(idx+1):02d}")
            name = find_str_val(item_dict, kw_name, f"منتج {idx+1}")
            category = find_str_val(item_dict, kw_cat, "عام" if lang == "AR" else "General")

            qty_sold = find_num_val(item_dict, kw_qty, 1.0)
            unit_price = find_num_val(item_dict, kw_price, 0.0)
            unit_cost = find_num_val(item_dict, kw_cost, 0.0)

            if unit_price > 100.0 and qty_sold > 1:
                unit_price = round(unit_price / qty_sold, 3)

            if unit_cost == 0.0 and unit_price > 0:
                unit_cost = round(unit_price * 0.65, 3)

            qty_wasted = find_num_val(item_dict, kw_wasted, 0.0)

            items.append({
                "code": code,
                "name": name,
                "category": category,
                "qty_sold": qty_sold,
                "unit_price": unit_price,
                "unit_cost": unit_cost,
                "qty_wasted": qty_wasted,
            })
    else:
        sample_cats = ["هواتف", "إكسسوارات", "مشروبات", "مواد غذائية", "مخبوزات"] if lang == "AR" else ["Phones", "Accessories", "Drinks", "Groceries", "Bakery"]
        sample_names = ["آيفون 15 بروماكس", "شاحن سريع 20 واط", "عصير برتقال طبيعي", "أرز بسمتي 5KG", "خبز توست أبيض"] if lang == "AR" else ["iPhone 15 Pro Max", "20W Fast Charger", "Natural Orange Juice", "Basmati Rice 5kg", "White Bread Pack"]

        for i in range(8):
            items.append({
                "code": f"SKU-10{i+1}",
                "name": sample_names[i % len(sample_names)],
                "category": sample_cats[i % len(sample_cats)],
                "qty_sold": (i + 1) * 5,
                "unit_price": (i + 1) * 12.5,
                "unit_cost": (i + 1) * 8.0,
                "qty_wasted": 1 if i % 3 == 0 else 0,
            })

    wb = openpyxl.Workbook()

    #  شيت موحد وملموم باسم "تقرير بصيرة"
    ws = wb.active
    ws.title = "تقرير بصيرة" if lang == "AR" else "Baseera Report"
    ws.sheet_view.rightToLeft = True if lang == "AR" else False
    ws.views.sheetView[0].showGridLines = True

    #  لوحة التنسيقات الفاخرة (Baseera Modern Luxury Palette)
    font_main_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_sub_title = Font(name="Segoe UI", size=10, italic=False, color="E2E8F0")
    font_section = Font(name="Segoe UI", size=11, bold=True, color="0F172A")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10, bold=False, color="1E293B")
    font_bold = Font(name="Segoe UI", size=10, bold=True, color="0F172A")

    font_kpi_title = Font(name="Segoe UI", size=9, bold=True, color="475569")
    font_kpi_val = Font(name="Segoe UI", size=14, bold=True, color="0F172A")
    font_kpi_val_emerald = Font(name="Segoe UI", size=14, bold=True, color="047857")
    font_footer = Font(name="Segoe UI", size=9.5, bold=True, color="1E3A8A")

    fill_title_banner = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Navy
    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Royal Blue
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Soft Slate
    fill_total = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid") # Soft Indigo
    fill_kpi = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_kpi_emerald = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    fill_footer = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid") # Light Indigo Footer

    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )
    kpi_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )
    total_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="medium", color="1E3A8A"),
        bottom=Side(style="double", color="1E3A8A")
    )

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    CURRENCY_FMT = f'#,##0.000 "{currency}"'
    PERCENT_FMT = '0.0%'
    INT_FMT = '#,##0'

    # =========================================================
    #  1. الهيدر الرئيسي فوق: "تقرير بصيرة"
    # =========================================================
    ws.merge_cells("A2:I2")
    ws.merge_cells("A3:I3")

    ws["A2"] = "تقرير بصيرة" if lang == "AR" else "Baseera Report"
    ws["A2"].font = font_main_title
    ws["A2"].fill = fill_title_banner
    ws["A2"].alignment = align_center

    report_date = datetime.now().strftime("%Y-%m-%d")
    sub_text = f"تقرير التحليل المالي والمبيعات والربحية - {company_name} | {report_date}" if lang == "AR" else f"Financial, Sales & Profitability Analysis - {company_name} | {report_date}"
    ws["A3"] = sub_text
    ws["A3"].font = font_sub_title
    ws["A3"].fill = fill_title_banner
    ws["A3"].alignment = align_center

    # تطبيق التعبئة والحدود على كامل خلايا الهيدر المدمجة
    for r_idx in [2, 3]:
        for c_idx in range(1, 10):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.fill = fill_title_banner

    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 18

    # =========================================================
    #  2. كروت المؤشرات الرئيسية (KPIs) - ملمومة ومرتبة
    # =========================================================
    # الأعمدة A-B (المبيعات), C-D (التكلفة), E-F (الهدر), G-I (صافي الربح)
    # أولاً نحسب مواضع الصفوف الديناميكية
    start_data_r = 9
    end_data_r = start_data_r + len(items) - 1
    tot_r = end_data_r + 1

    kpis = [
        ("إجمالي المبيعات" if lang == "AR" else "Total Revenue", f"=F{tot_r}", CURRENCY_FMT, fill_kpi, font_kpi_val, "A", "B"),
        ("التكلفة الإجمالية" if lang == "AR" else "Total Cost", f"=G{tot_r}", CURRENCY_FMT, fill_kpi, font_kpi_val, "C", "D"),
        ("إجمالي الهدر" if lang == "AR" else "Total Waste", f"=H{tot_r}", CURRENCY_FMT, fill_kpi, font_kpi_val, "E", "F"),
        ("صافي الربح" if lang == "AR" else "Net Profit", f"=I{tot_r}", CURRENCY_FMT, fill_kpi_emerald, font_kpi_val_emerald, "G", "I"),
    ]

    col_letter_map = {"A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7, "H": 8, "I": 9}

    for title, formula, fmt, fill_bg, val_font, c1, c2 in kpis:
        ws.merge_cells(f"{c1}5:{c2}5")
        ws.merge_cells(f"{c1}6:{c2}6")

        c_lbl = ws[f"{c1}5"]
        c_lbl.value = title
        c_lbl.font = font_kpi_title
        c_lbl.fill = fill_bg
        c_lbl.alignment = align_center

        c_val = ws[f"{c1}6"]
        c_val.value = formula
        c_val.font = val_font
        c_val.fill = fill_bg
        c_val.alignment = align_center
        c_val.number_format = fmt

        col_start_idx = col_letter_map[c1]
        col_end_idx = col_letter_map[c2]
        for r_i in [5, 6]:
            for col_i in range(col_start_idx, col_end_idx + 1):
                cell = ws.cell(row=r_i, column=col_i)
                cell.fill = fill_bg
                cell.border = kpi_border

    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 24

    # =========================================================
    #  3. جدول الأصناف والمبيعات الملموم (الجدول الرئيسي)
    # =========================================================
    headers = [
        "كود البند" if lang == "AR" else "SKU / Code",
        "اسم المنتج / السلعة" if lang == "AR" else "Product Name",
        "الفئة" if lang == "AR" else "Category",
        "الكمية" if lang == "AR" else "Qty Sold",
        "سعر الوحدة" if lang == "AR" else "Unit Price",
        "إجمالي المبيعات" if lang == "AR" else "Total Sales",
        "التكلفة الإجمالية" if lang == "AR" else "Total Cost",
        "خسارة الهدر" if lang == "AR" else "Waste Loss",
        "صافي الربح" if lang == "AR" else "Net Profit",
    ]

    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=8, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[8].height = 24

    for i, item in enumerate(items):
        r = start_data_r + i
        ws.cell(row=r, column=1, value=item.get("code", f"SKU-{i+1:02d}")).alignment = align_center
        ws.cell(row=r, column=2, value=item.get("name", "منتج")).alignment = align_right if lang == "AR" else align_left
        ws.cell(row=r, column=3, value=item.get("category", "عام")).alignment = align_center

        ws.cell(row=r, column=4, value=float(item.get("qty_sold", 0))).number_format = INT_FMT
        ws.cell(row=r, column=5, value=float(item.get("unit_price", 0.0))).number_format = CURRENCY_FMT

        # إجمالي المبيعات = الكمية * سعر الوحدة
        ws.cell(row=r, column=6, value=f"=D{r}*E{r}").number_format = CURRENCY_FMT

        # التكلفة الإجمالية = (الكمية المباعة + الكمية المهدرة) * تكلفة الوحدة
        unit_cost = float(item.get("unit_cost", 0.0))
        qty_wasted = float(item.get("qty_wasted", 0.0))
        ws.cell(row=r, column=7, value=f"=(D{r}+{qty_wasted})*{unit_cost}").number_format = CURRENCY_FMT

        # خسارة الهدر = الكمية المهدرة * تكلفة الوحدة
        ws.cell(row=r, column=8, value=qty_wasted * unit_cost).number_format = CURRENCY_FMT

        # صافي الربح = إجمالي المبيعات - التكلفة الإجمالية
        ws.cell(row=r, column=9, value=f"=F{r}-G{r}").number_format = CURRENCY_FMT

        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.font = font_data
            cell.border = thin_border
            if i % 2 == 1:
                cell.fill = fill_zebra
        ws.row_dimensions[r].height = 20

    # صف الإجمالي العام
    ws.cell(row=tot_r, column=1, value="الإجمالي العام" if lang == "AR" else "Grand Total").alignment = align_center
    ws.cell(row=tot_r, column=4, value=f"=SUM(D{start_data_r}:D{end_data_r})").number_format = INT_FMT
    ws.cell(row=tot_r, column=6, value=f"=SUM(F{start_data_r}:F{end_data_r})").number_format = CURRENCY_FMT
    ws.cell(row=tot_r, column=7, value=f"=SUM(G{start_data_r}:G{end_data_r})").number_format = CURRENCY_FMT
    ws.cell(row=tot_r, column=8, value=f"=SUM(H{start_data_r}:H{end_data_r})").number_format = CURRENCY_FMT
    ws.cell(row=tot_r, column=9, value=f"=SUM(I{start_data_r}:I{end_data_r})").number_format = CURRENCY_FMT

    for c in range(1, 10):
        cell = ws.cell(row=tot_r, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = total_border
    ws.row_dimensions[tot_r].height = 24

    # =========================================================
    #  4. ملخص الفئات المدمج وتحت التقرير
    # =========================================================
    cat_start_r = tot_r + 3
    ws.merge_cells(f"A{cat_start_r}:I{cat_start_r}")
    ws[f"A{cat_start_r}"] = "ملخص الأداء حسب الفئة" if lang == "AR" else "Performance Summary by Category"
    ws[f"A{cat_start_r}"].font = font_section
    ws[f"A{cat_start_r}"].alignment = align_right if lang == "AR" else align_left

    cat_hdr_r = cat_start_r + 1
    cat_headers = [
        "الفئة" if lang == "AR" else "Category",
        "إجمالي المبيعات" if lang == "AR" else "Total Sales",
        "صافي الربح" if lang == "AR" else "Net Profit",
        "نسبة المساهمة" if lang == "AR" else "Revenue Share %"
    ]

    cat_cols = [
        (1, 2), # A-B الفئة
        (3, 4), # C-D المبيعات
        (5, 6), # E-F الربح
        (7, 9), # G-I نسبة المساهمة
    ]

    for (c_s, c_e), h_text in zip(cat_cols, cat_headers):
        if c_s != c_e:
            ws.merge_cells(start_row=cat_hdr_r, start_column=c_s, end_row=cat_hdr_r, end_column=c_e)
        cell = ws.cell(row=cat_hdr_r, column=c_s, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

        for col_i in range(c_s, c_e + 1):
            ws.cell(row=cat_hdr_r, column=col_i).border = thin_border

    categories = list(dict.fromkeys([it["category"] for it in items if it.get("category")]))
    if not categories:
        categories = ["عام"] if lang == "AR" else ["General"]

    for idx, cat in enumerate(categories):
        r_i = cat_hdr_r + 1 + idx
        # دمج الخلايا للمظهر المرتّب
        for c_s, c_e in cat_cols:
            if c_s != c_e:
                ws.merge_cells(start_row=r_i, start_column=c_s, end_row=r_i, end_column=c_e)

        ws.cell(row=r_i, column=1, value=cat).alignment = align_center
        c_s_val = ws.cell(row=r_i, column=3, value=f"=SUMIF(C{start_data_r}:C{end_data_r}, A{r_i}, F{start_data_r}:F{end_data_r})")
        c_s_val.number_format = CURRENCY_FMT
        c_s_val.alignment = align_center

        c_p_val = ws.cell(row=r_i, column=5, value=f"=SUMIF(C{start_data_r}:C{end_data_r}, A{r_i}, I{start_data_r}:I{end_data_r})")
        c_p_val.number_format = CURRENCY_FMT
        c_p_val.alignment = align_center

        c_sh = ws.cell(row=r_i, column=7, value=f"=IF(F{tot_r}>0, C{r_i}/F{tot_r}, 0)")
        c_sh.number_format = PERCENT_FMT
        c_sh.alignment = align_center

        for col_c in range(1, 10):
            cell = ws.cell(row=r_i, column=col_c)
            cell.font = font_data
            cell.border = thin_border
            if idx % 2 == 1:
                cell.fill = fill_zebra

    cat_end_r = cat_hdr_r + len(categories)

    # =========================================================
    #  5. الفوتر السفلي الأنيق: "تقرير بصيرة" تحت (تلبية لطلب المستخدم)
    # =========================================================
    footer_r = cat_end_r + 3
    ws.merge_cells(f"A{footer_r}:I{footer_r}")

    footer_text = (
        f" تقرير بصيرة   |  تم إصدار هذا التقرير آلياً عبر منصة بصيرة للتحليلات الذكية - {company_name}"
        if lang == "AR"
        else f" Baseera Report   |  Generated automatically via Baseera Intelligence Platform - {company_name}"
    )

    ws[f"A{footer_r}"] = footer_text
    ws[f"A{footer_r}"].font = font_footer
    ws[f"A{footer_r}"].fill = fill_footer
    ws[f"A{footer_r}"].alignment = align_center

    for col_i in range(1, 10):
        ws.cell(row=footer_r, column=col_i).border = Border(
            top=Side(style="thin", color="C7D2FE"),
            bottom=Side(style="thin", color="C7D2FE")
        )

    ws.row_dimensions[footer_r].height = 28

    # =========================================================
    #  6. ضبط أبعاد الأعمدة تلقائياً ليكون ملموماً ومناسباً للشاشة
    # =========================================================
    col_widths = {
        "A": 16, # SKU
        "B": 24, # Product Name
        "C": 16, # Category
        "D": 12, # Qty
        "E": 16, # Unit Price
        "F": 18, # Total Sales
        "G": 18, # Total Cost
        "H": 16, # Waste Loss
        "I": 18, # Net Profit
    }

    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w

    wb.save(output_target)
    if hasattr(output_target, "seek"):
        output_target.seek(0)
    return output_target


def generate_dynamic_excel(payload: dict, output_target="تقرير_بصيرة.xlsx"):
    """دالة استقبال حمولة Make / Integromat مباشرة والتصدير بنفس الاسم."""
    return generate_baseera_excel(user_language="AR", client_data=payload, output_target=output_target)


def generate_universal_basira_report(raw_payload: dict, output_target="تقرير_بصيرة.xlsx"):
    """دالة التوافق الشامل."""
    lang = raw_payload.get("lang", "AR") if isinstance(raw_payload, dict) else "AR"
    return generate_baseera_excel(user_language=lang, client_data=raw_payload, output_target=output_target)


def generate_dynamic_basira_report(client_data: dict, output_target="تقرير_بصيرة.xlsx"):
    """دالة التوافق."""
    return generate_baseera_excel(user_language="AR", client_data=client_data, output_target=output_target)


def generate_ai_baseera_excel(excel_report_data: dict, output_target, company_name="", lang="AR"):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = excel_report_data.get("sheet_title", "تقرير بصيرة" if lang == "AR" else "Baseera Report")

    if lang == "AR":
        ws.sheet_view.rightToLeft = True

    # 1. Colors (Baseera Palette)
    color_bg_dark = "0B132B"
    color_bg_card = "1C2541"
    color_emerald = "10B981"
    color_text = "F3F4F6"
    color_muted = "9CA3AF"
    
    fill_header = PatternFill(start_color=color_bg_dark, end_color=color_bg_dark, fill_type="solid")
    fill_row = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_row_alt = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    font_title = Font(name="Segoe UI", size=20, bold=True, color=color_text)
    font_subtitle = Font(name="Segoe UI", size=12, italic=True, color=color_muted)
    font_header = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=11, color="1F2937")

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    # 2. Main Title Header
    ws.merge_cells("A1:E2")
    title_cell = ws["A1"]
    title_cell.value = "تقرير بصيرة للذكاء الاستراتيجي" if lang == "AR" else "Baseera Strategic Intelligence Report"
    title_cell.font = font_title
    title_cell.fill = fill_header
    title_cell.alignment = align_center

    ws.merge_cells("A3:E3")
    subtitle_cell = ws["A3"]
    subtitle_cell.value = company_name
    subtitle_cell.font = font_subtitle
    subtitle_cell.fill = PatternFill(start_color=color_bg_card, end_color=color_bg_card, fill_type="solid")
    subtitle_cell.alignment = align_center

    # 3. Table Headers
    headers = excel_report_data.get("headers", [])
    if not headers:
        headers = ["Metric", "Value", "Status", "Recommendation"]

    start_row = 5
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header_text)
        cell.font = font_header
        cell.fill = PatternFill(start_color=color_emerald, end_color=color_emerald, fill_type="solid")
        cell.alignment = align_center
        cell.border = thin_border

    # 4. Table Rows
    rows = excel_report_data.get("rows", [])
    current_row = start_row + 1
    
    for r_idx, row_data in enumerate(rows):
        for c_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=str(cell_value))
            cell.font = font_data
            cell.alignment = align_center if c_idx > 1 else (align_right if lang == "AR" else align_left)
            cell.border = thin_border
            cell.fill = fill_row_alt if r_idx % 2 == 1 else fill_row
        current_row += 1

    # 5. Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    wb.save(output_target)
    if hasattr(output_target, "seek"):
        output_target.seek(0)
    return output_target


if __name__ == "__main__":
    generate_baseera_excel(user_language="AR", output_filename="تقرير_بصيرة_عربي.xlsx")
    generate_baseera_excel(user_language="EN", output_filename="Baseera_Report_EN.xlsx")

